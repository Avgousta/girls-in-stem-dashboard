import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

xl = pd.read_excel('C:/Users/User/Downloads/Sage Girls in STEM - Mark Sheet.xlsx', sheet_name=None)

db_learners = [
    ('6e83c29e-5a64-4611-a89a-7ce57acc7212','LRN020','Aisha Saddiqa Hassim Hassim','UJ Academy SOS School'),
    ('3d88c1f0-d521-4270-a1e5-b08d67c1c700','LRN045','Aphiwe Mchiza','Diepdale Secondary School'),
    ('245afa12-855e-413e-8c13-456dfdbe4f5c','LRN021','Bushra Patel','UJ Academy SOS School'),
    ('02d25626-6d93-467c-9301-9fd0c6400a34','LRN044','Dakalo Mulaudzi','Diepdale Secondary School'),
    ('94898cd2-d4f7-4b91-8125-8a4ee38a37b8','LRN022','Dominique Rajanna','UJ Academy SOS School'),
    ('ec68d0e4-6ec1-4639-84c9-50a1f284826f','LRN023','Enhle Ngwane','UJ Academy SOS School'),
    ('f15c38b6-ab41-49a2-aa5c-599cf369d32e','LRN024','Fatima Davids','UJ Academy SOS School'),
    ('b4896527-03d1-46ac-81a1-2c4509f00744','LRN025','Hilwa Ibraahim','UJ Academy SOS School'),
    ('d379cc3a-4682-4476-87d2-d85fab0fc805','LRN026','Hirkumari Maisuriya','UJ Academy SOS School'),
    ('333412f0-34e7-48ea-b499-8a806c70fc74','LRN027','Husnaa Ishmial','UJ Academy SOS School'),
    ('30212ca7-92a7-47f4-8515-bf12299868a4','LRN043','Jasmine Anderson','Lancea Vale Secondary School'),
    ('6fb0bb25-3cb4-48d7-aa60-2012cf5c4435','LRN031','Sihaam Jeena','UJ Academy SOS School'),
    ('f1c9e5f0-d611-4a9e-9158-ea934024f413','LRN028','Kamogelo Simayile','UJ Academy SOS School'),
    ('1501d380-50f7-437b-8ceb-38949446b039','LRN042','Katlego Lesejane','Lancea Vale Secondary School'),
    ('e7977edb-d8c4-4d92-9c48-8212799f0cf6','LRN041','Leandra Bensen','Missouri Laan Secondary School'),
    ('945ce99b-b6bb-4192-85ba-9a14a12045b8','LRN040','Lebohang Hlangwane','Diepdale Secondary School'),
    ('683ca47b-90a0-4e18-bc8e-3e7ebf329132','LRN039','Lebogang Molapo','Missouri Laan Secondary School'),
    ('629520fa-5598-4f96-8dcd-f0531fd65bfa','LRN029','Mapula Mongalo','UJ Academy SOS School'),
    ('b590af7d-3cf7-4423-babc-61c9d19e0dc0','LRN038','Mukundi Magoma','Diepdale Secondary School'),
    ('0dae75cf-f17f-4f52-a6fb-8c91405b7514','LRN037','Ndodozo Netshiheni','Lancea Vale Secondary School'),
    ('dc93c1bf-5c5f-49e4-8a91-de0b506481e8','LRN036','Nokukhanya Mthethwa','Lancea Vale Secondary School'),
    ('f107eb9e-9000-4b0c-97a4-12b8fb758e6f','LRN035','Samantha Vukeya','Diepdale Secondary School'),
    ('9f1666cd-f571-4dba-b199-8f6330fd70f9','LRN030','Sameea Mangera','UJ Academy SOS School'),
    ('f8f05cd5-9365-45a3-8815-16fc52c1d038','LRN034','Shalati Baloyi','Diepdale Secondary School'),
    ('fa2e32f5-cdf0-4770-9ad9-46867abf6944','LRN033','Thandolwami Nyakeni','Missouri Laan Secondary School'),
    ('7ac79169-1ac6-4775-8e4a-b8adfd530543','LRN032','Yoliswa Ngxanga','Lancea Vale Secondary School'),
    ('76e7e26a-baae-479c-b4f7-22eb066fa770','LRN046','Anzani Ndou','Diepdale Secondary School'),
]

def norm(s):
    s = str(s).lower().strip()
    s = re.sub(r'[-\s]*(ujma|diepdale|missouri|lancea|sos|barnabas).*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[.\-_]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

db_map = {}
for lid, code, name, school in db_learners:
    db_map[norm(name)] = (lid, code, name, school)

ALIASES = {
    'aisha hasim':           'aisha saddiqa hassim hassim',
    'aisha saddiqa hassim':  'aisha saddiqa hassim hassim',
    'hilwa ebrahim':         'hilwa ibraahim',
    'hirkumair maisuriya':   'hirkumari maisuriya',
    'shalat baloyi':         'shalati baloyi',
    'mapula mangalo':        'mapula mongalo',
    'ndodzo netshiheni':     'ndodozo netshiheni',
    'lebogang hlangwane':    'lebohang hlangwane',
    'lebohang molapo':       'lebogang molapo',
}

def lookup(raw_name):
    n = norm(raw_name)
    if n in db_map: return db_map[n]
    if n in ALIASES: return db_map.get(ALIASES[n])
    parts = n.split()
    if len(parts) >= 2:
        for k, v in db_map.items():
            kp = k.split()
            if parts[0] in kp and parts[-1] in kp:
                return v
    return None

# Collect all learners across the 3 main sheets
seen = {}  # norm_name -> {sheets, school, match}
SHEETS = [('Grade9_2024', 2024), ('Grade10_2025', 2025), ('Grade11_2026', 2026)]

for sheet_name, year in SHEETS:
    df = xl[sheet_name]
    for _, row in df.iterrows():
        raw = str(row.get('Learner', '')).strip()
        if not raw or raw == 'nan': continue
        school_raw = str(row.get('School', '')).strip()
        key = norm(raw)
        if key not in seen:
            seen[key] = {'raw': raw, 'school': school_raw, 'sheets': [], 'match': lookup(raw)}
        seen[key]['sheets'].append(year)

# ── Build workbook ──────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Learner Matching Report'

# Styles
PURPLE      = 'FF4F2D7F'
LIGHT_PURP  = 'FFEDE9FE'
GREEN       = 'FF16A34A'
LIGHT_GREEN = 'FFDCFCE7'
AMBER       = 'FFD97706'
LIGHT_AMBER = 'FFFEF3C7'
WHITE       = 'FFFFFFFF'
GRAY_LIGHT  = 'FFF8F8F8'
BORDER_COL  = 'FFE2E8F0'

def cell_style(ws, row, col, value, bold=False, bg=None, fg='FF000000', wrap=False, align='left', sz=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    thin = Side(style='thin', color=BORDER_COL)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

# ── TITLE ───────────────────────────────────────────────────────────────
ws.merge_cells('A1:H1')
title = ws.cell(row=1, column=1, value='Girls in STEM — Learner Matching Report')
title.font = Font(name='Arial', bold=True, size=16, color=WHITE)
title.fill = PatternFill('solid', start_color=PURPLE)
title.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:H2')
sub = ws.cell(row=2, column=1, value='Review matched learners and fill in Learner Codes for unmatched names so historic marks can be imported.')
sub.font = Font(name='Arial', size=10, color='FF4B5563')
sub.alignment = Alignment(horizontal='center', vertical='center')
sub.fill = PatternFill('solid', start_color=LIGHT_PURP)
ws.row_dimensions[2].height = 20

# ── LEGEND ──────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 6  # spacer

ws.merge_cells('A4:H4')
leg = ws.cell(row=4, column=1, value='Legend:   ✅ Green = Matched automatically   |   ⚠ Amber = Needs your input — enter the Learner Code in column D')
leg.font = Font(name='Arial', size=10, bold=True, color='FF374151')
leg.alignment = Alignment(horizontal='left', vertical='center')
leg.fill = PatternFill('solid', start_color='FFFFF7ED')
ws.row_dimensions[4].height = 18

ws.row_dimensions[5].height = 6  # spacer

# ── HEADER ROW ──────────────────────────────────────────────────────────
headers = [
    ('A', 'Name in Spreadsheet',   22),
    ('B', 'School (Spreadsheet)',   22),
    ('C', 'Sheets Found In',        20),
    ('D', 'Learner Code ← FILL IN', 20),
    ('E', 'Matched DB Name',        28),
    ('F', 'School (Database)',       28),
    ('G', 'Status',                  16),
    ('H', 'Notes',                   30),
]
for i, (col_letter, header, width) in enumerate(headers, 1):
    c = ws.cell(row=6, column=i, value=header)
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', start_color='FF1F2937')
    c.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='FF374151')
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[6].height = 22

# ── SECTION: MATCHED ────────────────────────────────────────────────────
row = 7
ws.merge_cells(f'A{row}:H{row}')
sec = ws.cell(row=row, column=1, value=f'  ✅  MATCHED — {sum(1 for v in seen.values() if v["match"])} learners automatically matched')
sec.font = Font(name='Arial', bold=True, size=11, color='FF166534')
sec.fill = PatternFill('solid', start_color='FFF0FDF4')
sec.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[row].height = 22
row += 1

for key, info in sorted(seen.items(), key=lambda x: x[1]['raw']):
    if not info['match']:
        continue
    lid, code, db_name, db_school = info['match']
    sheets_str = ', '.join(str(y) for y in sorted(set(info['sheets'])))
    note = ''
    if norm(info['raw']) != norm(db_name):
        note = f'Name variation matched: "{info["raw"]}" → "{db_name}"'

    cell_style(ws, row, 1, info['raw'],  bold=True,  bg=LIGHT_GREEN, fg='FF14532D')
    cell_style(ws, row, 2, info['school'],            bg=LIGHT_GREEN, fg='FF166534')
    cell_style(ws, row, 3, sheets_str,   align='center', bg=LIGHT_GREEN, fg='FF166534')
    cell_style(ws, row, 4, code,         bold=True, align='center', bg=LIGHT_GREEN, fg='FF15803D')
    cell_style(ws, row, 5, db_name,                   bg=LIGHT_GREEN, fg='FF14532D')
    cell_style(ws, row, 6, db_school,                 bg=LIGHT_GREEN, fg='FF166534')
    cell_style(ws, row, 7, '✅ Matched', bold=True, align='center', bg=LIGHT_GREEN, fg='FF166534')
    cell_style(ws, row, 8, note,                      bg=LIGHT_GREEN, fg='FF4D7C5E', wrap=True)
    ws.row_dimensions[row].height = 18
    row += 1

# ── SPACER ───────────────────────────────────────────────────────────────
row += 1

# ── SECTION: UNMATCHED ──────────────────────────────────────────────────
unmatched_items = [(k, v) for k, v in sorted(seen.items(), key=lambda x: x[1]['raw']) if not v['match']]
ws.merge_cells(f'A{row}:H{row}')
sec2 = ws.cell(row=row, column=1, value=f'  ⚠  NEEDS YOUR INPUT — {len(unmatched_items)} learners not found in the database — enter Learner Code in column D')
sec2.font = Font(name='Arial', bold=True, size=11, color='FF92400E')
sec2.fill = PatternFill('solid', start_color='FFFEF3C7')
sec2.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[row].height = 22
row += 1

instruction = ws.cell(row=row, column=4, value='▼ TYPE LEARNER CODE HERE (e.g. LRN047)')
instruction.font = Font(name='Arial', bold=True, size=9, color='FFB45309', italic=True)
instruction.fill = PatternFill('solid', start_color='FFFEF9C3')
instruction.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 16
row += 1

for key, info in unmatched_items:
    sheets_str = ', '.join(str(y) for y in sorted(set(info['sheets'])))
    cell_style(ws, row, 1, info['raw'],  bold=True,  bg=LIGHT_AMBER, fg='FF92400E')
    cell_style(ws, row, 2, info['school'],            bg=LIGHT_AMBER, fg='FFB45309')
    cell_style(ws, row, 3, sheets_str, align='center', bg=LIGHT_AMBER, fg='FFB45309')
    # Column D = editable input cell (styled differently)
    input_cell = ws.cell(row=row, column=4, value='')
    input_cell.font = Font(name='Arial', bold=True, size=11, color=PURPLE)
    input_cell.fill = PatternFill('solid', start_color='FFFFFFED')
    input_cell.alignment = Alignment(horizontal='center', vertical='center')
    thick = Side(style='medium', color='FF7C3AED')
    input_cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    cell_style(ws, row, 5, '—',  align='center', bg=LIGHT_AMBER, fg='FFD1D5DB')
    cell_style(ws, row, 6, '—',  align='center', bg=LIGHT_AMBER, fg='FFD1D5DB')
    cell_style(ws, row, 7, '⚠ Not Found', bold=True, align='center', bg=LIGHT_AMBER, fg='FFB45309')
    cell_style(ws, row, 8, 'Enter learner code to include their marks in the import', bg=LIGHT_AMBER, fg='FFB45309', wrap=True)
    ws.row_dimensions[row].height = 20
    row += 1

# ── FOOTER ───────────────────────────────────────────────────────────────
row += 1
ws.merge_cells(f'A{row}:H{row}')
footer = ws.cell(row=row, column=1, value='Once you have filled in the learner codes above, save this file and share it back — I will run the import automatically.')
footer.font = Font(name='Arial', size=10, bold=True, color='FF4F2D7F')
footer.fill = PatternFill('solid', start_color=LIGHT_PURP)
footer.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 22

# ── SECOND SHEET: existing DB learners for reference ────────────────────
ws2 = wb.create_sheet('Database Learners (Reference)')
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 32

ref_headers = ['Learner Code', 'Full Name', 'Grade', 'School']
for i, h in enumerate(ref_headers, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = PatternFill('solid', start_color='FF1F2937')
    c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 20

all_db = [
    ('LRN001','Nombuso Mkhize',10,'UJ Academy SOS School'),
    ('LRN002','Habza Ali',10,'UJ Academy SOS School'),
    ('LRN003','Kamono Mofokeng',10,'St Barnabas SOS School'),
    ('LRN004','Kayla Sonickson',10,'St Barnabas SOS School'),
    ('LRN005','Maaknun Ali',10,'UJ Academy SOS School'),
    ('LRN006','Maimoona Pandor',10,'St Barnabas SOS School'),
    ('LRN007','Mammudi Chosane',10,'UJ Academy SOS School'),
    ('LRN008','Mohau Mukazi',10,'UJ Academy SOS School'),
    ('LRN009','Mosa Leeu',10,'St Barnabas SOS School'),
    ('LRN010','Paballo Matsapola',10,'UJ Academy SOS School'),
    ('LRN011','Reabiloe Motsile',10,'UJ Academy SOS School'),
    ('LRN012','Sihaam Shire',10,'UJ Academy SOS School'),
    ('LRN013','Tayleigh Morgan',10,'St Barnabas SOS School'),
    ('LRN014','Tlotliso Buang',10,'St Barnabas SOS School'),
    ('LRN015','Jennipher Amungwat',10,'St Barnabas SOS School'),
    ('LRN016','Danika Wood',10,'UJ Academy SOS School'),
    ('LRN017','Jordyn Sinclair',10,'St Barnabas SOS School'),
    ('LRN018','Nsovo Makama',10,'St Barnabas SOS School'),
    ('LRN019','Ndzetelo Ngobeni',10,'UJ Academy SOS School'),
    ('LRN020','Aisha Saddiqa Hassim Hassim',11,'UJ Academy SOS School'),
    ('LRN021','Bushra Patel',11,'UJ Academy SOS School'),
    ('LRN022','Dominique Rajanna',11,'UJ Academy SOS School'),
    ('LRN023','Enhle Ngwane',11,'UJ Academy SOS School'),
    ('LRN024','Fatima Davids',11,'UJ Academy SOS School'),
    ('LRN025','Hilwa Ibraahim',11,'UJ Academy SOS School'),
    ('LRN026','Hirkumari Maisuriya',11,'UJ Academy SOS School'),
    ('LRN027','Husnaa Ishmial',11,'UJ Academy SOS School'),
    ('LRN028','Kamogelo Simayile',11,'UJ Academy SOS School'),
    ('LRN029','Mapula Mongalo',11,'UJ Academy SOS School'),
    ('LRN030','Sameea Mangera',11,'UJ Academy SOS School'),
    ('LRN031','Sihaam Jeena',11,'UJ Academy SOS School'),
    ('LRN032','Yoliswa Ngxanga',11,'Lancea Vale Secondary School'),
    ('LRN033','Thandolwami Nyakeni',11,'Missouri Laan Secondary School'),
    ('LRN034','Shalati Baloyi',11,'Diepdale Secondary School'),
    ('LRN035','Samantha Vukeya',11,'Diepdale Secondary School'),
    ('LRN036','Nokukhanya Mthethwa',11,'Lancea Vale Secondary School'),
    ('LRN037','Ndodozo Netshiheni',11,'Lancea Vale Secondary School'),
    ('LRN038','Mukundi Magoma',11,'Diepdale Secondary School'),
    ('LRN039','Lebogang Molapo',11,'Missouri Laan Secondary School'),
    ('LRN040','Lebohang Hlangwane',11,'Diepdale Secondary School'),
    ('LRN041','Leandra Bensen',11,'Missouri Laan Secondary School'),
    ('LRN042','Katlego Lesejane',11,'Lancea Vale Secondary School'),
    ('LRN043','Jasmine Anderson',11,'Lancea Vale Secondary School'),
    ('LRN044','Dakalo Mulaudzi',11,'Diepdale Secondary School'),
    ('LRN045','Aphiwe Mchiza',11,'Diepdale Secondary School'),
    ('LRN046','Anzani Ndou',11,'Diepdale Secondary School'),
]

for i, (code, name, grade, school) in enumerate(all_db, 2):
    bg = LIGHT_PURP if grade == 11 else 'FFF8FAFC'
    for j, val in enumerate([code, name, f'Grade {grade}', school], 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = Font(name='Arial', size=10, bold=(j==1))
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left' if j!=1 else 'center', vertical='center')
        thin = Side(style='thin', color=BORDER_COL)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws2.row_dimensions[i].height = 16

ws.freeze_panes = 'A7'
out = 'C:/Users/User/Downloads/GirlsSTEM_Learner_Matching_Report.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Matched: {sum(1 for v in seen.values() if v["match"])}')
print(f'Unmatched: {len(unmatched_items)}')
print(f'Total learners: {len(seen)}')
